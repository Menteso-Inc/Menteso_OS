"""
PCT Agent — Patent Cooperation Treaty
Reads WIPO resultList.xls → scrapes patent pages → extracts contacts → outputs Work Report Excel.

Input:  resultList.xls (from WIPO PatentScope weekly browse) or .xlsx with same columns
Output: Work Report DD-Month-YYYY.xlsx matching the standard work report format
"""
import os
import re
import time
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

from shared.memory import load_memory, save_learning, get_best_strategy
from shared.self_debug import run_with_self_debug
from .scraper import download_wipo_excel
from .browser import PatentBrowser, BrowserStopRequested
from .pdf_extractor import extract_contacts_from_pdf
from .pipeline import ChunkedPipelineManager, ProgressFile
from .tests import tests

# Pipeline config
# Lowered threshold so even small runs benefit from chunked processing.
PIPELINE_THRESHOLD = 20
DEFAULT_CHUNK_SIZE = 500
# Short cooldown between chunks (was 1.5s) — captcha solver handles bursts.
CHUNK_COOLDOWN_SECONDS = 0.3

# Parallel pipeline worker counts. Conservative defaults to keep WIPO happy:
# 4 simultaneous browsers is enough for ~50-150 rows/min, low enough that
# WIPO rarely rate-limits a single home IP. Push higher only if you have
# IP rotation (Tor or paid proxies).
DEFAULT_BROWSER_WORKERS = 4
DEFAULT_DOWNLOAD_WORKERS = 6
DEFAULT_OCR_WORKERS = 3

# Fast-mode level table. Each level maps to a runtime profile.
# Level 1 (😇) is the safe sequential baseline; level 5 (😈) maxes out
# parallelism. Worker counts in the fast-mode levels are TOTAL across
# all parallel pipelines — ChunkedPipelineManager divides them down per
# chunk pipeline at construction time.
# All levels run HEADLESS. The captcha solver handles DIY OCR + GPT-4o
# Vision in the headless worker silently — no Chromium window appears.
# Only when both auto-tiers fail does manual_fallback open a SEPARATE
# visible popup Chromium for the human to solve, which closes itself
# the moment the captcha clears.
FAST_MODE_LEVELS = {
    1: {"emoji": "😇", "label": "Safe",       "fast": False, "pipelines": 1, "browsers": 1, "downloads": 1,  "ocr": 1, "headless": True},
    2: {"emoji": "🙂", "label": "Mild",       "fast": True,  "pipelines": 1, "browsers": 2, "downloads": 3,  "ocr": 2, "headless": True},
    3: {"emoji": "😏", "label": "Balanced",   "fast": True,  "pipelines": 2, "browsers": 4, "downloads": 6,  "ocr": 3, "headless": True},
    4: {"emoji": "😎", "label": "Aggressive", "fast": True,  "pipelines": 3, "browsers": 6, "downloads": 9,  "ocr": 4, "headless": True},
    5: {"emoji": "😈", "label": "Max",        "fast": True,  "pipelines": 4, "browsers": 8, "downloads": 12, "ocr": 6, "headless": True},
}
# Start in Balanced (L3) by default — parallel pool with conservative
# concurrency. Users can drop to L1 Safe via the slider if WIPO throttles,
# or push higher (L4/L5) when they have IP rotation in place.
FAST_MODE_DEFAULT_LEVEL = 3


def _resolve_headless(fast_profile):
    """Decide headless mode for this run.
    Precedence: explicit PCT_HEADLESS_MODE env var (true/false) > profile.headless.
    """
    env_raw = (os.getenv("PCT_HEADLESS_MODE") or "").strip().lower()
    if env_raw in ("1", "true", "yes"):
        return True
    if env_raw in ("0", "false", "no"):
        return False
    if isinstance(fast_profile, dict):
        return fast_profile.get("headless", True)
    return True


def resolve_fast_level(input_data):
    """Pick the active fast-mode level for this run.

    Precedence: input_data['fast_level'] (from UI) > env PCT_FAST_LEVEL >
    legacy env PCT_FAST_MODE (true → 3, false → 1) > default level 1.
    Returns a (level_int, profile_dict) tuple.
    """
    level = None
    if isinstance(input_data, dict):
        raw = input_data.get("fast_level")
        if raw is not None:
            try:
                level = int(raw)
            except (TypeError, ValueError):
                level = None
    if level is None:
        env_level = os.getenv("PCT_FAST_LEVEL")
        if env_level is not None:
            try:
                level = int(env_level)
            except ValueError:
                level = None
    if level is None:
        legacy = (os.getenv("PCT_FAST_MODE") or "").lower()
        if legacy in ("1", "true", "yes"):
            level = 3
        elif legacy in ("0", "false", "no"):
            level = 1
    if level is None:
        level = FAST_MODE_DEFAULT_LEVEL
    level = max(1, min(5, level))
    return level, FAST_MODE_LEVELS[level]

AGENT_CONFIG = {
    "name": "PCT Agent",
    "description": (
        "Patent Cooperation Treaty agent. Reads WIPO resultList Excel, "
        "scrapes PatentScope patent pages, downloads RO/101 or 306 PDFs, "
        "extracts email/phone/name contacts, and outputs a Work Report Excel."
    ),
    "role": "Patent Data Processor",
    "goal": "Extract contact information from WIPO patent filings",
    "status": "active",
    "version": "2.0.0",
    "requires_llm": False,
    "accepts_upload": True,
    "upload_types": [".xlsx", ".xls"],
    "input_fields": [
        {
            "name": "mode",
            "type": "select",
            "label": "Input Mode",
            "options": ["Upload Excel", "Download from WIPO"],
        },
    ],
    "sub_agents": ["Scraper", "PDF Extractor", "CAPTCHA Solver"],
}

# Decorative pause between major lifecycle steps. Set to 0 in production for
# maximum throughput; bump to ~0.3 if you want a more visible step-by-step
# log for demos.
STEP_DELAY = 0


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def id_to_url(patent_id):
    """Convert patent ID to WIPO PatentScope URL.
    e.g. 'WO/2025/097187' → 'https://patentscope.wipo.int/search/en/WO2025097187'
    """
    clean = patent_id.replace("/", "")
    return f"https://patentscope.wipo.int/search/en/{clean}"


def extract_country(appl_no):
    """Extract country code from application number.
    e.g. 'US2023/078371' → 'US', 'AT2024/060361' → 'AT'
    """
    match = re.match(r'^([A-Z]{2})', str(appl_no).strip())
    return match.group(1) if match else ""


# ---------------------------------------------------------------------------
# Excel reader — supports both .xls (xlrd) and .xlsx (openpyxl)
# ---------------------------------------------------------------------------
def read_input_excel(file_path, on_step=None):
    """Read WIPO resultList Excel. Returns list of row dicts.
    Handles:
      - .xls files (xlrd) — WIPO default download format
      - .xlsx files (openpyxl)
      - Skips blank rows and gazette header rows
    """
    ext = Path(file_path).suffix.lower()

    if ext == ".xls":
        if not HAS_XLRD:
            raise ImportError("xlrd is required for .xls files — pip install xlrd")
        return _read_xls(file_path, on_step)
    elif ext in (".xlsx", ".xlsm"):
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is required for .xlsx files — pip install openpyxl")
        return _read_xlsx(file_path, on_step)
    else:
        raise ValueError(f"Unsupported file format: {ext}")


def _read_xls(file_path, on_step=None):
    """Read .xls file using xlrd (WIPO resultList format)."""
    wb = xlrd.open_workbook(file_path)
    ws = wb.sheet_by_index(0)

    if on_step:
        on_step(f"[Excel Reader] Sheet: '{ws.name}' — {ws.nrows} rows x {ws.ncols} cols")

    # Find header row — look for a row where first cell is "ID"
    header_row = None
    for r in range(min(10, ws.nrows)):
        val = str(ws.cell_value(r, 0)).strip()
        if val.upper() == "ID":
            header_row = r
            break

    if header_row is None:
        raise ValueError("Could not find header row with 'ID' column in the Excel file")

    headers = [str(ws.cell_value(header_row, c)).strip() for c in range(ws.ncols)]
    if on_step:
        on_step(f"[Excel Reader] Headers found at row {header_row + 1}: {headers}")

    col_map = _map_columns(headers)

    rows = []
    for r in range(header_row + 1, ws.nrows):
        row_vals = [ws.cell_value(r, c) for c in range(ws.ncols)]
        patent_id = str(row_vals[col_map.get("id", 0)]).strip()
        if not patent_id:
            continue

        rows.append({
            "id": patent_id,
            "title": str(row_vals[col_map.get("title", 1)]).strip(),
            "appl_no": str(row_vals[col_map.get("appl_no", 3)]).strip(),
            "applicant": str(row_vals[col_map.get("applicant", 5)]).strip(),
            "kind": str(row_vals[col_map.get("kind", 2)]).strip() if "kind" in col_map else "",
            "ipc": str(row_vals[col_map.get("ipc", 4)]).strip() if "ipc" in col_map else "",
        })

    if on_step:
        on_step(f"[Excel Reader] Parsed {len(rows)} patent entries")

    return rows


def _read_xlsx(file_path, on_step=None):
    """Read .xlsx file using openpyxl."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    if on_step:
        on_step(f"[Excel Reader] Sheet: '{ws.title}' — {ws.max_row} rows x {ws.max_column} cols")

    # Find header row
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=False):
        for cell in row:
            if cell.value and str(cell.value).strip().upper() == "ID":
                header_row = cell.row
                break
        if header_row:
            break

    if header_row is None:
        raise ValueError("Could not find header row with 'ID' column in the Excel file")

    header_cells = list(ws.iter_rows(min_row=header_row, max_row=header_row))[0]
    headers = [str(cell.value or "").strip() for cell in header_cells]
    if on_step:
        on_step(f"[Excel Reader] Headers found at row {header_row}: {headers}")

    col_map = _map_columns(headers)

    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        row_vals = list(row)
        patent_id = str(row_vals[col_map.get("id", 0)] or "").strip()
        if not patent_id:
            continue

        rows.append({
            "id": patent_id,
            "title": str(row_vals[col_map.get("title", 1)] or "").strip(),
            "appl_no": str(row_vals[col_map.get("appl_no", 3)] or "").strip(),
            "applicant": str(row_vals[col_map.get("applicant", 5)] or "").strip(),
            "kind": str(row_vals[col_map.get("kind", 2)] or "").strip() if "kind" in col_map else "",
            "ipc": str(row_vals[col_map.get("ipc", 4)] or "").strip() if "ipc" in col_map else "",
        })

    if on_step:
        on_step(f"[Excel Reader] Parsed {len(rows)} patent entries")

    wb.close()
    return rows


def _map_columns(headers):
    """Map header names to column indices."""
    col_map = {}
    for i, h in enumerate(headers):
        h_upper = h.upper()
        if h_upper == "ID":
            col_map["id"] = i
        elif h_upper == "TITLE":
            col_map["title"] = i
        elif h_upper == "APPLICANT":
            col_map["applicant"] = i
        elif "APPL" in h_upper:
            # Must come AFTER "APPLICANT" check — "Appl.No" contains "APPL"
            col_map["appl_no"] = i
        elif h_upper == "IPC":
            col_map["ipc"] = i
        elif h_upper == "KIND":
            col_map["kind"] = i
    return col_map


def _stop_requested(input_data):
    checker = (input_data or {}).get("stop_requested")
    if callable(checker):
        return checker
    return lambda: False


def _register_stop_handler(input_data, handler):
    registrar = (input_data or {}).get("register_stop_handler")
    if callable(registrar) and callable(handler):
        registrar(handler)


def _build_partial_result(status, results, total, found_count, not_found_count,
                          error_count, output_path="", execution_time=0, tests_result=None):
    return {
        "status": status,
        "results": results,
        "summary": {
            "total": total,
            "processed": len(results),
            "found": found_count,
            "not_found": not_found_count,
            "errors": error_count,
        },
        "output_file": output_path,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "execution_time": round(execution_time, 2),
        "attempts": 1,
        "tests": tests_result or {},
    }


# ---------------------------------------------------------------------------
# Main agent runner
# ---------------------------------------------------------------------------
def run_agent(input_data=None, on_step=None):
    """
    Execute the PCT agent.
    input_data expects:
      - file_path: path to uploaded Excel (.xls or .xlsx)
      - mode: "upload" or "wipo_download"
    """
    start_time = time.time()
    agent_name = "pct_agent"

    def step(msg):
        if on_step:
            on_step(msg)

    def browser_event(event_data):
        if on_step:
            on_step({"type": "browser", **event_data})

    # --- Step 1: Load memory ---
    step("Loading PCT agent memory...")
    time.sleep(STEP_DELAY)
    memory = load_memory(agent_name)
    runs = memory["stats"]["total_runs"]
    rate = memory["stats"]["success_rate"]
    step(f"Memory loaded — {runs} past runs, {rate:.0%} success rate")
    time.sleep(STEP_DELAY)

    # --- Step 2: Select strategy ---
    step("Selecting best strategy...")
    strategy = get_best_strategy(agent_name, "process_excel", default="sequential_scrape")
    step(f"Strategy: {strategy}")
    time.sleep(STEP_DELAY)

    # --- Step 3: Validate input ---
    if not input_data:
        input_data = {}

    mode = input_data.get("mode", "upload")
    file_path = input_data.get("file_path")
    gazette = input_data.get("gazette")
    stop_requested = _stop_requested(input_data)

    if mode == "wipo_download":
        gazette_msg = f" for week {gazette}" if gazette else ""
        step(f"[Mode: WIPO Download] Downloading Excel from PatentScope{gazette_msg}...")
        time.sleep(STEP_DELAY)

        def do_download():
            return download_wipo_excel(gazette=gazette, on_step=on_step)

        dl_result = run_with_self_debug(do_download, max_retries=2, on_step=on_step)
        if dl_result["status"] == "success" and dl_result["result"]:
            file_path = dl_result["result"]
            step(f"Excel downloaded: {file_path}")
        else:
            step("[WIPO Download] Could not download Excel — check network/URL")
            save_learning(agent_name, "process_excel", "failure",
                          "WIPO download failed", "try_different_headers")
            return _failure("Could not download Excel from WIPO PatentScope")

    if not file_path or not os.path.exists(file_path):
        step(f"ERROR: Excel file not found: {file_path}")
        return _failure(f"Excel file not found: {file_path}")

    # --- Step 4: Read input Excel ---
    step(f"[Excel Reader] Opening: {Path(file_path).name}")
    time.sleep(STEP_DELAY)

    try:
        patent_rows = read_input_excel(file_path, on_step=step)
    except Exception as e:
        step(f"ERROR: Failed to read Excel: {e}")
        return _failure(f"Failed to read Excel: {e}")

    if not patent_rows:
        step("ERROR: No patent entries found in the Excel file")
        return _failure("No patent entries found in the Excel file")

    step(f"Ready to process {len(patent_rows)} patent entries")
    time.sleep(STEP_DELAY)

    if stop_requested():
        step("Stop requested before processing started")
        return _build_partial_result("stopped", [], len(patent_rows), 0, 0, 0)

    # --- Step 5: Choose mode — pipeline (large) or sequential (small) ---
    # fast_level (1-5) controls the runtime profile. Level 1 is the safe
    # chunked-sequential path (one browser). Levels 2-5 go through the
    # parallel ChunkedPipelineManager with progressively more workers.
    fast_level, fast_profile = resolve_fast_level(input_data)
    step(
        f"[Mode] fast_level={fast_level} {fast_profile['emoji']} {fast_profile['label']} "
        f"→ pipelines={fast_profile['pipelines']}, browsers={fast_profile['browsers']}, "
        f"downloads={fast_profile['downloads']}, ocr={fast_profile['ocr']}"
    )
    # Make the live level reader available to dispatcher inner loops so they
    # can react to mid-run changes at chunk boundaries.
    live_level_getter = (input_data or {}).get("get_live_fast_level")
    if len(patent_rows) >= PIPELINE_THRESHOLD:
        if fast_profile["fast"]:
            return _run_pipeline_mode(
                patent_rows, file_path, agent_name, strategy,
                start_time, step, browser_event, stop_requested,
                (input_data or {}).get("register_stop_handler"),
                fast_profile=fast_profile,
                live_level_getter=live_level_getter,
            )
        return _run_chunked_sequential_mode(
            patent_rows, file_path, agent_name, strategy,
            start_time, step, browser_event, stop_requested,
            (input_data or {}).get("register_stop_handler"),
            fast_profile=fast_profile,
            live_level_getter=live_level_getter,
        )

    # --- Sequential mode (< PIPELINE_THRESHOLD rows) ---
    step("Launching browser for WIPO scraping (headless)...")
    step("A Chromium popup will only appear if DIY OCR + GPT-4o Vision both fail to solve a captcha.")
    time.sleep(STEP_DELAY)

    results = []
    found_count = 0
    not_found_count = 0
    error_count = 0
    total = len(patent_rows)

    # All levels run headless by default. Only the manual_fallback tier
    # opens a separate visible Chromium popup when needed.
    headless_mode = _resolve_headless(fast_profile)
    patent_browser = PatentBrowser(
        headless=headless_mode,
        on_step=on_step,
        stop_requested=stop_requested,
    )
    _register_stop_handler(input_data, patent_browser.force_stop)
    try:
        patent_browser.start()
    except Exception as e:
        step(f"ERROR: Could not launch browser: {e}")
        return _failure(f"Browser launch failed: {e}")

    aborted_early = False
    try:
        for idx, row_data in enumerate(patent_rows, start=1):
            if stop_requested():
                step(f"Stop requested - finishing with {len(results)} processed row(s)")
                aborted_early = True
                break

            patent_id = row_data["id"]
            title = row_data["title"]
            url = id_to_url(patent_id)
            country = extract_country(row_data["appl_no"])
            doc_id = patent_id.replace("/", "_")

            step(f"[Row {idx}/{total}] Processing: {patent_id}")

            # Dashboard browser preview: navigate
            browser_event({
                "event": "navigate",
                "url": url,
                "row": idx,
                "total": total,
                "patent_id": patent_id,
                "title": title,
                "applicant": row_data["applicant"],
                "country": country,
            })

            # Per-row try/except: any unexpected exception is captured as an
            # error result and the loop continues — one bad row never kills
            # the whole run. BrowserStopRequested is the one exception we
            # honor by breaking out cleanly.
            try:
                step(f"[Row {idx}] [Browser] Opening patent page & searching for RO/101 PDF...")
                pdf_path = patent_browser.scrape_patent(url, doc_id, on_step=on_step)

                if not pdf_path:
                    browser_event({
                        "event": "no_pdf",
                        "url": url,
                        "row": idx,
                        "total": total,
                        "patent_id": patent_id,
                    })
                    step(f"[Row {idx}] No RO/101 PDF found")
                    results.append(_row_result(
                        idx, row_data, url, country, "not_found",
                        reason="No RO/101 PDF found on patent page",
                    ))
                    not_found_count += 1
                    continue

                # PDF downloaded — extract contacts
                browser_event({
                    "event": "extracting",
                    "url": url,
                    "row": idx,
                    "total": total,
                    "patent_id": patent_id,
                })
                step(f"[Row {idx}] [PDF Extractor] Extracting contacts...")
                contacts = extract_contacts_from_pdf(pdf_path, on_step=on_step)

                emails = contacts.get("emails", [])
                phones = contacts.get("phones", [])
                name = contacts.get("name", "")
                status = contacts["status"]

                browser_event({
                    "event": "contacts",
                    "url": url,
                    "row": idx,
                    "total": total,
                    "patent_id": patent_id,
                    "title": title,
                    "emails": emails,
                    "phones": phones,
                    "name": name,
                    "status": status,
                    "found_count": found_count + (1 if status == "found" else 0),
                    "not_found_count": not_found_count + (1 if status == "not_found" else 0),
                    "error_count": error_count,
                })

                if status == "found":
                    found_count += 1
                    step(
                        f"[Row {idx}] FOUND: "
                        f"{', '.join(emails[:2]) if emails else 'no email'} | "
                        f"{', '.join(phones[:2]) if phones else 'no phone'}"
                    )
                elif status == "not_found":
                    not_found_count += 1
                    step(f"[Row {idx}] No contact info found in PDF")
                else:
                    error_count += 1
                    step(f"[Row {idx}] Error: {contacts.get('error', 'Unknown')}")

                results.append(_row_result(
                    idx, row_data, url, country, status,
                    emails=emails, phones=phones, name=name,
                ))
            except BrowserStopRequested:
                step(f"[Row {idx}] Stop requested - ending run immediately and saving partial output")
                aborted_early = True
                break
            except Exception as row_exc:
                step(f"[Row {idx}] UNEXPECTED error: {row_exc} — recording and moving on")
                error_count += 1
                results.append(_row_result(
                    idx, row_data, url, country, "error",
                    reason=f"row_exception: {str(row_exc)[:300]}",
                ))

            # No per-row sleep — pacing is enforced upstream by RequestPacer
            # at every page.goto, and the captcha solver absorbs any burst.

    finally:
        step("[Browser] Closing browser...")
        try:
            patent_browser.close()
        except Exception as e:
            step(f"[Browser] Close raised (ignored): {e}")

    # --- Step 6: Generate Work Report Excel ---
    step("Generating Work Report Excel...")
    time.sleep(STEP_DELAY)

    run_status = "stopped" if (stop_requested() or aborted_early) else "success"
    output_path = ""
    if results:
        try:
            output_path = generate_work_report(results, on_step=step)
            step(f"Output saved: {Path(output_path).name}")
        except Exception as e:
            step(f"[Output] FAILED to write Work Report: {e} — "
                 f"{len(results)} rows are still in memory but the .xlsx was not produced")
    else:
        step("No processed rows available to write into Work Report")
    time.sleep(STEP_DELAY)

    # --- Step 7: Self-test ---
    step("Running self-tests...")
    time.sleep(STEP_DELAY)

    agent_result = _build_partial_result(
        run_status, results, total, found_count, not_found_count,
        error_count, output_path=output_path,
    )
    agent_result["summary"]["skipped"] = len([r for r in results if r["status"] == "skipped"])

    test_result = tests.run(agent_result)
    agent_result["tests"] = test_result

    if test_result["passed"]:
        step(f"All {test_result['total']} self-tests passed!")
    else:
        step(f"Self-tests: {test_result['passed_count']}/{test_result['total']} passed")

    # --- Step 8: Save learning ---
    execution_time = time.time() - start_time
    insight = (
        f"Processed {total} rows: {found_count} contacts found, "
        f"{not_found_count} not found, {error_count} errors"
    )
    save_learning(
        agent_name, "process_excel",
        run_status if run_status == "stopped" else ("success" if found_count > 0 or not_found_count > 0 else "partial"),
        insight, strategy, execution_time,
    )

    step(f"Learning saved. Total time: {execution_time:.1f}s")
    time.sleep(STEP_DELAY)

    if run_status == "stopped":
        step(
            f"STOPPED - partial output created with {len(results)} processed rows "
            f"({found_count} found, {not_found_count} not found, {error_count} errors)"
        )
    else:
        step(
            f"DONE - {found_count} contacts found, {not_found_count} not found, "
            f"{error_count} errors out of {total} rows"
        )

    agent_result["execution_time"] = round(execution_time, 2)
    agent_result["attempts"] = 1
    return agent_result

    step(
        f"DONE — {found_count} contacts found, {not_found_count} not found, "
        f"{error_count} errors out of {total} rows"
    )

    agent_result["execution_time"] = round(execution_time, 2)
    agent_result["attempts"] = 1
    return agent_result


# ---------------------------------------------------------------------------
# Pipeline mode — parallel processing for large datasets
# ---------------------------------------------------------------------------
def _run_pipeline_mode(patent_rows, file_path, agent_name, strategy,
                       start_time, step, browser_event, stop_requested=lambda: False,
                       register_stop_handler=None, fast_profile=None,
                       live_level_getter=None):
    """Run the parallel pipeline for 50+ rows.  No artificial delays."""
    profile = fast_profile or FAST_MODE_LEVELS[3]
    browser_workers = profile["browsers"]
    download_workers = profile["downloads"]
    ocr_workers = profile["ocr"]
    parallel_pipelines = profile["pipelines"]

    step(f"[Pipeline] Large dataset ({len(patent_rows)} rows) — parallel pipeline mode")
    step(
        f"[Pipeline] {browser_workers} browsers + {download_workers} downloaders + "
        f"{ocr_workers} OCR across {parallel_pipelines} pipeline(s)"
    )

    # Check for resume
    resume_path = ProgressFile.find_latest(file_path)
    if resume_path:
        completed = ProgressFile.load_completed(resume_path)
        step(f"[Pipeline] Resuming — {len(completed)}/{len(patent_rows)} already done")
    else:
        resume_path = None

    # Honor profile's headless setting; PCT_HEADLESS_MODE env can override.
    headless_mode = _resolve_headless(profile)

    # Pool warm-up: re-use the pacer's "boost after captcha" mechanism to
    # slow down the first ~5 requests right after pool launch. This avoids
    # smashing WIPO with a burst of N parallel requests immediately after a
    # level upgrade (L1 → L5), which the site reads as a bot signature
    # and reacts to by serving empty/throttled pages — leading to a wave
    # of false not_found results.
    try:
        from shared.pacing import default_pacer as _pacer
        if _pacer is not None and browser_workers >= 2:
            _pacer.report_captcha()  # trips the post-captcha multiplier
            step(
                f"[Pipeline] Warm-up: throttling the first ~{_pacer.boost_requests} "
                f"requests so WIPO sees a gradual ramp, not a burst"
            )
    except Exception:
        pass

    # Build and run pipeline. Wrap in try/except so even if the parallel
    # pipeline dies catastrophically we still write whatever rows finished.
    pipeline = ChunkedPipelineManager(
        patent_rows=patent_rows,
        on_step=step,
        browser_workers=browser_workers,
        download_workers=download_workers,
        ocr_workers=ocr_workers,
        parallel_pipelines=parallel_pipelines,
        headless=headless_mode,
        resume_path=resume_path,
        input_file=file_path,
        live_level_getter=live_level_getter,
    )

    results = []
    pipeline_error = None
    try:
        results = pipeline.run()
    except Exception as e:
        pipeline_error = e
        step(f"[Pipeline] ERROR during run: {e} — saving whatever we have")

    # Generate Work Report from whatever rows survived.
    output_path = ""
    if results:
        step("Generating Work Report Excel...")
        try:
            output_path = generate_work_report(results, on_step=step)
            step(f"Output saved: {Path(output_path).name}")
        except Exception as e:
            step(f"[Output] Could not write Work Report: {e} — "
                 f"{len(results)} rows are in the JSONL log")
    else:
        step("No rows processed — Work Report not generated")

    # Self-tests
    step("Running self-tests...")
    found_count = sum(1 for r in results if r["status"] == "found")
    not_found_count = sum(1 for r in results if r["status"] == "not_found")
    error_count = sum(1 for r in results if r["status"] not in ("found", "not_found"))
    total = len(patent_rows)

    agent_result = {
        "status": "success",
        "results": results,
        "summary": {
            "total": total,
            "processed": len(results),
            "found": found_count,
            "not_found": not_found_count,
            "errors": error_count,
        },
        "output_file": output_path,
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }

    test_result = tests.run(agent_result)
    agent_result["tests"] = test_result

    if test_result["passed"]:
        step(f"All {test_result['total']} self-tests passed!")
    else:
        step(f"Self-tests: {test_result['passed_count']}/{test_result['total']} passed")

    # Save learning
    execution_time = time.time() - start_time
    insight = (
        f"Pipeline processed {total} rows: {found_count} contacts found, "
        f"{not_found_count} not found, {error_count} errors"
    )
    save_learning(agent_name, "process_excel", "success" if found_count > 0 else "partial",
                  insight, strategy, execution_time)

    step(f"Learning saved. Total time: {execution_time:.1f}s")
    step(f"DONE — {found_count} contacts found, {not_found_count} not found, "
         f"{error_count} errors out of {total} rows")

    agent_result["execution_time"] = round(execution_time, 2)
    agent_result["attempts"] = 1
    return agent_result


# ---------------------------------------------------------------------------
# Output generator — Work Report format
# ---------------------------------------------------------------------------
def _run_chunked_sequential_mode(patent_rows, file_path, agent_name, strategy,
                       start_time, step, browser_event, stop_requested=lambda: False,
                       register_stop_handler=None, fast_profile=None,
                       live_level_getter=None):
    """Run large datasets in stable sequential chunks.

    Single browser, one row at a time, but split into chunks for stability.
    Slower than _run_pipeline_mode but uses only one connection to WIPO so
    captcha rate is minimal. Used as the safe fallback when fast_level <= 1.

    `live_level_getter`, if provided, is consulted at chunk boundaries so a
    user dragging the slider up to fast-mode causes future chunks to be
    handed off to _run_pipeline_mode instead of being processed here.
    """
    profile = fast_profile or FAST_MODE_LEVELS[1]
    total = len(patent_rows)
    total_chunks = max(1, (total + DEFAULT_CHUNK_SIZE - 1) // DEFAULT_CHUNK_SIZE)

    step(f"[Pipeline] Large dataset ({total} rows) - sequential chunk mode")
    step(f"[Pipeline] Splitting into {total_chunks} chunk(s) of up to {DEFAULT_CHUNK_SIZE} rows")
    step("[Pipeline] Processing chunks strictly one by one for stability")

    for idx, row in enumerate(patent_rows, start=1):
        row["_row_idx"] = idx

    results = []
    processed_rows = set()
    found_count = 0
    not_found_count = 0
    error_count = 0

    step("Launching browser for chunked WIPO scraping...")
    # All levels run headless. A visible Chromium popup only opens if BOTH
    # DIY OCR and GPT-4o Vision fail to solve a captcha (manual_fallback).
    headless_mode = _resolve_headless(profile)
    if headless_mode:
        step(
            f"[Mode {profile['emoji']} {profile['label']}] Running headless — "
            f"only opens a Chromium popup if DIY OCR + GPT-4o Vision both fail."
        )
    else:
        step(f"[Mode {profile['emoji']} {profile['label']}] Headed mode (PCT_HEADLESS_MODE=false) — visible Chromium window will open.")

    patent_browser = PatentBrowser(
        headless=headless_mode,
        on_step=step,
        stop_requested=stop_requested,
    )
    if callable(register_stop_handler):
        register_stop_handler(patent_browser.force_stop)
    try:
        patent_browser.start()
    except Exception as e:
        step(f"ERROR: Could not launch browser: {e}")
        return _failure(f"Browser launch failed: {e}")

    upgrade_remaining_rows = None

    def _check_live_upgrade(current_row_index):
        """If the user dragged the slider to a fast level, return the
        slice of rows we should hand off to the parallel pipeline. Called
        between rows so switching takes effect within seconds, not chunks.
        """
        if not callable(live_level_getter):
            return None
        try:
            live = live_level_getter()
        except Exception:
            return None
        if not live or live not in FAST_MODE_LEVELS:
            return None
        if not FAST_MODE_LEVELS[live]["fast"]:
            return None
        # Bail with everything from current row onwards.
        remaining = patent_rows[current_row_index:]
        new_profile = FAST_MODE_LEVELS[live]
        step(
            f"[Mode] Live upgrade to L{live} {new_profile['emoji']} {new_profile['label']} "
            f"— handing off {len(remaining)} remaining row(s) to the parallel pipeline"
        )
        return remaining

    try:
        for chunk_index in range(total_chunks):
            if stop_requested():
                step(f"[Chunk {chunk_index + 1}/{total_chunks}] Stop requested before next chunk")
                break

            chunk_start = chunk_index * DEFAULT_CHUNK_SIZE

            # Pre-chunk live-level check — catches the case where the user
            # upgraded between chunks (including before the very first one).
            upgrade_remaining_rows = _check_live_upgrade(chunk_start)
            if upgrade_remaining_rows:
                break

            chunk_rows = patent_rows[chunk_start:chunk_start + DEFAULT_CHUNK_SIZE]
            if not chunk_rows:
                continue

            first_row = chunk_rows[0]["_row_idx"]
            last_row = chunk_rows[-1]["_row_idx"]
            step(
                f"[Chunk {chunk_index + 1}/{total_chunks}] Starting rows "
                f"{first_row}-{last_row} ({len(chunk_rows)} rows)"
            )

            chunk_results = 0
            for row_pos, row_data in enumerate(chunk_rows):
                if stop_requested():
                    step(
                        f"[Chunk {chunk_index + 1}/{total_chunks}] Stop requested - "
                        f"finishing with {len(results)} processed row(s)"
                    )
                    break

                # Per-row live-level check — this is what makes the slider
                # feel instant. If user upgraded mid-chunk, bail right now
                # with whatever rows are still unprocessed.
                upgrade_remaining_rows = _check_live_upgrade(chunk_start + row_pos)
                if upgrade_remaining_rows:
                    break

                row_no = row_data["_row_idx"]
                if row_no in processed_rows:
                    step(f"[Chunk {chunk_index + 1}/{total_chunks}] Skipping duplicate row {row_no}")
                    continue

                patent_id = row_data["id"]
                title = row_data["title"]
                url = id_to_url(patent_id)
                country = extract_country(row_data["appl_no"])
                doc_id = patent_id.replace("/", "_")

                step(f"[Row {row_no}/{total}] Processing: {patent_id}")

                browser_event({
                    "event": "navigate",
                    "url": url,
                    "row": row_no,
                    "total": total,
                    "patent_id": patent_id,
                    "title": title,
                    "applicant": row_data["applicant"],
                    "country": country,
                })

                # Per-row try/except: an unexpected exception becomes an
                # error row, not a crash. Stop requests still break out
                # cleanly. This is what keeps the PCT agent running through
                # any single-row failure.
                try:
                    step(f"[Row {row_no}] [Browser] Opening patent page & searching for RO/101 PDF...")
                    pdf_path = patent_browser.scrape_patent(url, doc_id, on_step=step)

                    if not pdf_path:
                        browser_event({
                            "event": "no_pdf",
                            "url": url,
                            "row": row_no,
                            "total": total,
                            "patent_id": patent_id,
                        })
                        step(f"[Row {row_no}] No RO/101 PDF found")
                        results.append(_row_result(
                            row_no, row_data, url, country, "not_found",
                            reason="No RO/101 PDF found on patent page",
                        ))
                        processed_rows.add(row_no)
                        not_found_count += 1
                        chunk_results += 1
                        continue

                    browser_event({
                        "event": "extracting",
                        "url": url,
                        "row": row_no,
                        "total": total,
                        "patent_id": patent_id,
                    })
                    step(f"[Row {row_no}] [PDF Extractor] Extracting contacts...")
                    contacts = extract_contacts_from_pdf(pdf_path, on_step=step)

                    emails = contacts.get("emails", [])
                    phones = contacts.get("phones", [])
                    name = contacts.get("name", "")
                    status = contacts["status"]

                    if status == "found":
                        found_count += 1
                        step(
                            f"[Row {row_no}] FOUND: "
                            f"{', '.join(emails[:2]) if emails else 'no email'} | "
                            f"{', '.join(phones[:2]) if phones else 'no phone'}"
                        )
                    elif status == "not_found":
                        not_found_count += 1
                        step(f"[Row {row_no}] No contact info found in PDF")
                    else:
                        error_count += 1
                        step(f"[Row {row_no}] Error: {contacts.get('error', 'Unknown')}")

                    browser_event({
                        "event": "contacts",
                        "url": url,
                        "row": row_no,
                        "total": total,
                        "patent_id": patent_id,
                        "title": title,
                        "emails": emails,
                        "phones": phones,
                        "name": name,
                        "status": status,
                        "found_count": found_count,
                        "not_found_count": not_found_count,
                        "error_count": error_count,
                    })

                    results.append(_row_result(
                        row_no, row_data, url, country, status,
                        emails=emails, phones=phones, name=name,
                    ))
                    processed_rows.add(row_no)
                    chunk_results += 1
                except BrowserStopRequested:
                    step(
                        f"[Chunk {chunk_index + 1}/{total_chunks}] Stop requested - "
                        "ending run immediately and saving partial output"
                    )
                    break
                except Exception as row_exc:
                    step(f"[Row {row_no}] UNEXPECTED error: {row_exc} — recording and moving on")
                    error_count += 1
                    results.append(_row_result(
                        row_no, row_data, url, country, "error",
                        reason=f"row_exception: {str(row_exc)[:300]}",
                    ))
                    processed_rows.add(row_no)
                    chunk_results += 1

                # No per-row sleep — RequestPacer handles backoff naturally.

            step(
                f"[Chunk {chunk_index + 1}/{total_chunks}] Finished rows "
                f"{first_row}-{last_row} - appended {chunk_results} result row(s); "
                f"combined dataset now has {len(results)} row(s)"
            )

            if stop_requested():
                break

            # If the inner row-loop broke due to a live upgrade, propagate.
            if upgrade_remaining_rows:
                break

            if chunk_index < total_chunks - 1:
                step(
                    f"[Chunk {chunk_index + 1}/{total_chunks}] Cooling down "
                    f"{CHUNK_COOLDOWN_SECONDS:.1f}s before next chunk"
                )
                time.sleep(CHUNK_COOLDOWN_SECONDS)
    finally:
        step("[Browser] Closing browser...")
        try:
            patent_browser.close()
        except Exception as e:
            step(f"[Browser] Close raised (ignored): {e}")

    # If the user upgraded to fast mode mid-run, hand the remaining rows
    # off to the parallel pipeline now. Its results merge with what we
    # already produced sequentially.
    if upgrade_remaining_rows and not stop_requested():
        try:
            live = live_level_getter() if callable(live_level_getter) else None
            new_profile = FAST_MODE_LEVELS.get(live or 3, FAST_MODE_LEVELS[3])
            pipeline_result = _run_pipeline_mode(
                upgrade_remaining_rows, file_path, agent_name, strategy,
                start_time, step, browser_event, stop_requested,
                register_stop_handler,
                fast_profile=new_profile,
                live_level_getter=live_level_getter,
            )
            # Merge pipeline results with our sequential ones
            pipeline_rows = (pipeline_result or {}).get("results", []) or []
            seen_rows = {r.get("row") for r in results}
            for r in pipeline_rows:
                if r.get("row") not in seen_rows:
                    results.append(r)
                    seen_rows.add(r.get("row"))
                    if r.get("status") == "found":
                        found_count += 1
                    elif r.get("status") == "not_found":
                        not_found_count += 1
                    else:
                        error_count += 1
        except Exception as e:
            step(f"[Mode] Pipeline handoff failed: {e} — saving sequential results only")

    results = sorted(results, key=lambda item: item.get("row", 0))
    if len(results) != total:
        step(
            f"[Pipeline] WARNING: final result count {len(results)} does not match "
            f"input rows {total}"
        )

    step("Generating Work Report Excel...")
    run_status = "stopped" if stop_requested() else "success"
    output_path = ""
    if results:
        try:
            output_path = generate_work_report(results, on_step=step)
            step(f"Output saved: {Path(output_path).name}")
        except Exception as e:
            step(f"[Output] FAILED to write Work Report: {e} — "
                 f"results were already streamed to the resumable JSONL log")
    else:
        step("No processed rows available to write into Work Report")

    step("Running self-tests...")
    agent_result = _build_partial_result(
        run_status, results, total, found_count, not_found_count,
        error_count, output_path=output_path,
    )

    test_result = tests.run(agent_result)
    agent_result["tests"] = test_result

    if test_result["passed"]:
        step(f"All {test_result['total']} self-tests passed!")
    else:
        step(f"Self-tests: {test_result['passed_count']}/{test_result['total']} passed")

    execution_time = time.time() - start_time
    insight = (
        f"Sequential chunk processing completed {total} rows: "
        f"{found_count} found, {not_found_count} not found, {error_count} errors"
    )
    save_learning(
        agent_name, "process_excel",
        run_status if run_status == "stopped" else ("success" if found_count > 0 or not_found_count > 0 else "partial"),
        insight, strategy, execution_time,
    )

    step(f"Learning saved. Total time: {execution_time:.1f}s")
    if run_status == "stopped":
        step(
            f"STOPPED - partial output created with {len(results)} processed rows "
            f"({found_count} found, {not_found_count} not found, {error_count} errors)"
        )
    else:
        step(
            f"DONE - {found_count} contacts found, {not_found_count} not found, "
            f"{error_count} errors out of {total} rows"
        )

    agent_result["execution_time"] = round(execution_time, 2)
    agent_result["attempts"] = 1
    return agent_result


WORK_REPORT_HEADERS = [
    "Publication Number", "Title", "Application No", "Applicant",
    "Url", "Cat", "Phone No", "Email", "Name", "Country",
    "Date", "Researcher", "Deadline",
]


def generate_work_report(results, on_step=None):
    """Generate a Work Report Excel matching the standard output format."""
    results = sorted(results, key=lambda item: item.get("row", 0))
    wb = openpyxl.Workbook()
    ws = wb.active

    # Sheet name: "Work Report DD-Month-YYYY"
    date_str = datetime.now().strftime("%d-%B-%Y")
    sheet_name = f"Work Report {date_str}"
    ws.title = sheet_name[:31]

    # Write headers with bold font
    for col, header in enumerate(WORK_REPORT_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    run_date = f"Shared - {datetime.now().strftime('%d-%m-%Y')}"

    for i, r in enumerate(results, start=2):
        ws.cell(row=i, column=1, value=r.get("patent_id", ""))
        ws.cell(row=i, column=2, value=r.get("title", ""))
        ws.cell(row=i, column=3, value=r.get("appl_no", ""))
        ws.cell(row=i, column=4, value=r.get("applicant", ""))
        ws.cell(row=i, column=5, value=r.get("url", ""))
        ws.cell(row=i, column=6, value="")             # Cat — filled manually
        ws.cell(row=i, column=7, value="; ".join(r.get("phones", [])))
        ws.cell(row=i, column=8, value="; ".join(r.get("emails", [])))
        ws.cell(row=i, column=9, value=r.get("name", ""))
        ws.cell(row=i, column=10, value=r.get("country", ""))
        ws.cell(row=i, column=11, value=run_date)
        ws.cell(row=i, column=12, value="")            # Researcher — filled manually
        ws.cell(row=i, column=13, value="")            # Deadline — filled manually

    # Save to outputs/
    output_dir = Path(__file__).parent.parent.parent / "outputs"
    output_dir.mkdir(exist_ok=True)

    output_name = f"Work Report {date_str}.xlsx"
    output_path = output_dir / output_name
    counter = 2
    while output_path.exists():
        output_name = f"Work Report {date_str} book{counter}.xlsx"
        output_path = output_dir / output_name
        counter += 1

    wb.save(str(output_path))

    if on_step:
        on_step(f"[Output] Work Report saved: {output_name} ({len(results)} rows)")

    return str(output_path)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _row_result(idx, row_data, url, country, status,
                emails=None, phones=None, name="", reason=""):
    """Build a standardised row result dict."""
    return {
        "row": idx,
        "patent_id": row_data["id"],
        "title": row_data["title"],
        "appl_no": row_data["appl_no"],
        "applicant": row_data["applicant"],
        "url": url,
        "country": country,
        "status": status,
        "emails": emails or [],
        "phones": phones or [],
        "name": name,
        "reason": reason,
    }


def _failure(error_msg):
    """Build a standardised failure result."""
    return {
        "status": "failure",
        "error": error_msg,
        "results": [],
        "summary": {"total": 0, "found": 0, "not_found": 0, "errors": 0},
    }
